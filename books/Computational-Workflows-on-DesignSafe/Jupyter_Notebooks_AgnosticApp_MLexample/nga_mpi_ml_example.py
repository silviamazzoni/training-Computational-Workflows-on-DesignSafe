#!/usr/bin/env python3
"""
NGA flatfile (.xlsx) + HDF5 waveform processing with MPI4Py, plus a simple ML regression demo.

UPDATED per your rules:
- Flatfile filenames are ignored entirely.
- Use ONLY NGAFilename_H1 and NGAFilename_H2 from the filename-mapping CSV.
- Skip V for now.
- Do NOT skip an RSN if some other file is missing; only require H1 and H2.
  -> If H1 missing OR H2 missing -> skip that RSN (log it).
- Robust dataset lookup:
  Mapping may store dataset stems without ".AT2"; we try common variants.

Author: Silvia Mazzoni (silviamazzoni@yahoo.com)
"""

from __future__ import annotations

import argparse
import csv
import math
import os
from dataclasses import dataclass
from typing import Dict, List, Optional, Sequence, Tuple

import numpy as np
import h5py
from mpi4py import MPI
from openpyxl import load_workbook

from pathlib import Path

from dataclasses import field

import json
from dataclasses import fields

from typing import Union

PathLike = Union[str, Path]


def _to_float_nan999(x: Optional[str], default: float = np.nan) -> float:
    v = _to_float(x, default=default)
    if np.isfinite(v) and v <= -900.0:   # catches -999, -999.0, etc.
        return np.nan
    return v



def resolve_outdir(outdir: PathLike) -> Path:
    """
    Resolve + create an output directory.
    Safe to call on rank 0 (MPI) or in serial scripts.

    Returns an absolute Path.
    """
    p = Path(outdir).expanduser().resolve()
    p.mkdir(parents=True, exist_ok=True)
    return p

def outpath(outdir: PathLike, prefix: str, stem: str, ext: str = ".csv") -> Path:
    """
    Build an output path like:
      <outdir>/<prefix>_<stem><ext>

    Examples:
      outpath("out_process", "NGAWest2", "metrics_H1") -> out_process/NGAWest2_metrics_H1.csv
      outpath(outdir, "NGAWest2", "summary_plots", ".pdf") -> .../NGAWest2_summary_plots.pdf
    """
    od = Path(outdir)
    return od / f"{prefix}_{stem}{ext}"



# ---------------------------
# RotD calcs
# 1) Add this RotD computation block

@dataclass
class RotDRow:
    rsn: str
    rotd: str              # RotD0, RotD50, RotD100
    angle_deg: float       # angle at which percentile was attained (nearest)
    pga: float             # max abs accel (g or units of your timeseries)
    amp_range: float       # max - min (same units as accel)
    dt_peaks: float        # |tmax - tmin|  (seconds) if dt known else NaN
    dt_peaks_norm: float   # dt_peaks/duration if dt known else NaN
    duration: float        # (npts-1)*dt if dt known else NaN
    n_angles: int
    ok: int
    err: str



def compute_rotd_summaries(
    a1: np.ndarray,
    a2: np.ndarray,
    dt: float,
    angles_deg: np.ndarray,
) -> List[RotDRow]:
    """
    Compute RotD0/50/100 for a few time-domain scalars:
      - pga (max abs accel)
      - amp_range (max - min)
      - dt_peaks, dt_peaks_norm using (argmax, argmin) of the *governing* rotated component
        where governing means the one with larger max-abs at that angle.

    Notes:
      - dt may be NaN; in that case time-based metrics are NaN.
      - This is a time-domain RotD variant. It mirrors the RotD idea (min/median/max over rotations).
    """
    try:
        if a1.shape != a2.shape:
            raise ValueError(f"H1/H2 length mismatch: {a1.shape} vs {a2.shape}")

        n = a1.size
        if n < 2:
            raise ValueError("Timeseries too short")

        # Precompute trig
        theta = np.deg2rad(angles_deg.astype(float))
        c = np.cos(theta)[:, None]   # (nang,1)
        s = np.sin(theta)[:, None]   # (nang,1)

        # Rotate into orthogonal pair for each angle: u, v (nang, npts)
        u = c * a1[None, :] + s * a2[None, :]
        v = -s * a1[None, :] + c * a2[None, :]

        # Max-abs per angle for each component
        u_maxabs = np.max(np.abs(u), axis=1)
        v_maxabs = np.max(np.abs(v), axis=1)

        # Governing scalar per angle (like PGA RotD concept)
        pga_theta = np.maximum(u_maxabs, v_maxabs)

        # For amp_range, compute range for governing component at each angle
        # Governing component = u if u_maxabs >= v_maxabs else v
        use_u = (u_maxabs >= v_maxabs)
        # Compute max/min for both then select
        u_max = np.max(u, axis=1); u_min = np.min(u, axis=1)
        v_max = np.max(v, axis=1); v_min = np.min(v, axis=1)
        amp_range_theta = np.where(use_u, u_max - u_min, v_max - v_min)

        # Time-based metrics for governing component
        if np.isfinite(dt) and dt > 0:
            # indices of max/min for u and v
            u_imax = np.argmax(u, axis=1); u_imin = np.argmin(u, axis=1)
            v_imax = np.argmax(v, axis=1); v_imin = np.argmin(v, axis=1)

            imax = np.where(use_u, u_imax, v_imax)
            imin = np.where(use_u, u_imin, v_imin)

            tmax = imax.astype(float) * dt
            tmin = imin.astype(float) * dt
            dt_peaks_theta = np.abs(tmax - tmin)
            duration = float((n - 1) * dt)
            dt_peaks_norm_theta = dt_peaks_theta / duration if duration > 0 else np.nan
        else:
            dt_peaks_theta = np.full_like(pga_theta, np.nan, dtype=float)
            dt_peaks_norm_theta = np.full_like(pga_theta, np.nan, dtype=float)
            duration = np.nan

        # Percentiles over angles
        def _pick_rotd(name: str, q: float) -> RotDRow:
            # Compute percentile value
            val = np.quantile(pga_theta, q, method="linear")

            # Choose a representative angle (closest pga_theta to percentile)
            k = int(np.argmin(np.abs(pga_theta - val)))
            return RotDRow(
                rsn="",
                rotd=name,
                angle_deg=float(angles_deg[k]),
                pga=float(pga_theta[k]),
                amp_range=float(amp_range_theta[k]),
                dt_peaks=float(dt_peaks_theta[k]),
                dt_peaks_norm=float(dt_peaks_norm_theta[k]),
                duration=float(duration),
                n_angles=int(len(angles_deg)),
                ok=1,
                err="",
            )

        rows = [
            _pick_rotd("RotD0", 0.0),
            _pick_rotd("RotD50", 0.5),
            _pick_rotd("RotD100", 1.0),
        ]
        return rows

    except Exception as e:
        return [RotDRow(
            rsn="",
            rotd="RotD50",
            angle_deg=np.nan,
            pga=np.nan,
            amp_range=np.nan,
            dt_peaks=np.nan,
            dt_peaks_norm=np.nan,
            duration=np.nan,
            n_angles=int(len(angles_deg)),
            ok=0,
            err=str(e),
        )]

# 2) Add a CSV writer for RotD rows
def write_rotd_csv(path: str, rows: List[RotDRow]) -> None:
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    fieldnames = list(RotDRow.__annotations__.keys())
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in rows:
            w.writerow({k: getattr(r, k) for k in fieldnames})

# ---------------------------
# Helpers
# ---------------------------

def _get_dt_from_dataset(ds: h5py.Dataset) -> float:
    """
    Read dt from dataset attribute 'dtHeader'. Returns np.nan if missing/unparseable.
    Handles cases where attrs are stored as scalar, 1-element array, bytes, or string.
    """
    if ds is None:
        return np.nan

    if "dtHeader" not in ds.attrs:
        return np.nan

    v = ds.attrs.get("dtHeader")

    try:
        # h5py can return numpy scalars/arrays
        if isinstance(v, np.ndarray):
            if v.size == 0:
                return np.nan
            v = v.reshape(-1)[0]

        # bytes -> str
        if isinstance(v, (bytes, bytearray)):
            v = v.decode("utf-8", errors="ignore")

        # string like "0.005" -> float
        return float(v)
    except Exception:
        return np.nan


def _norm_key(s: str) -> str:
    return (s or "").strip().lower()


def _pick_col_value(row: Dict[str, str], candidates: Sequence[str], required: bool = False) -> Optional[str]:
    keymap = {_norm_key(k): k for k in row.keys()}
    for c in candidates:
        k = _norm_key(c)
        if k in keymap:
            val = (row.get(keymap[k]) or "").strip()
            if val != "":
                return val
    if required:
        raise KeyError(
            f"Missing required column. Tried: {list(candidates)}. "
            f"Available: {list(row.keys())[:50]} ..."
        )
    return None


def _to_float(x: Optional[str], default: float = np.nan) -> float:
    if x is None:
        return default
    try:
        return float(x)
    except Exception:
        return default


def _is_missing_token(x: str) -> bool:
    s = x.strip()
    return s == "" or s.lower() in {"nan", "none"} or s in {"-999", "-999.0"}


def _ensure_file_exists(path: str, label: str) -> None:
    if not os.path.exists(path):
        raise FileNotFoundError(f"{label} does not exist: {path}")
    if not os.path.isfile(path):
        raise FileNotFoundError(f"{label} is not a file: {path}")


def load_csv_rows(path: str) -> List[Dict[str, str]]:
    with open(path, "r", newline="", encoding="utf-8", errors="ignore") as f:
        rdr = csv.DictReader(f)
        rows = [r for r in rdr]
    if len(rows) == 0:
        raise ValueError(f"CSV is empty: {path}")
    return rows


# ---------------------------
# Read flatfile XLSX (NO filenames used)
# ---------------------------

# FLATFILE_RSN_COLS = ["Record Sequence Number", "RSN", "NGARSN"]
# FLATFILE_MW_COLS = ["Earthquake Magnitude", "Mw", "Magnitude", "mag", "M"]
# FLATFILE_VS30_COLS = ["Vs30 (m/s) selected for analysis", "Vs30", "vs30"]
# FLATFILE_EpiD_COLS = ["EpiD (km)"]
# FLATFILE_Rjb_COLS = ["Joyner-Boore Dist. (km)"]
# FLATFILE_HypD_COLS = ["HypD (km)"]
# FLATFILE_ClstD_COLS = ["ClstD (km)"]
# FLATFILE_PGA_COLS = ["PGA (g)"]
# FLATFILE_PGV_COLS = ["PGV (cm/sec)"]
# FLATFILE_DT_COLS = ["DT", "dt", "DeltaT"]


FLATFILE_COLS = {}
FLATFILE_COLS['RSN'] = ["Record Sequence Number", "RSN", "NGARSN"]
FLATFILE_COLS['MW'] = ["Earthquake Magnitude", "Mw", "Magnitude", "mag", "M"]
FLATFILE_COLS['VS30'] = ["Vs30 (m/s) selected for analysis", "Vs30", "vs30"]
FLATFILE_COLS['EpiD'] = ["EpiD (km)"]
FLATFILE_COLS['Rjb'] = ["Joyner-Boore Dist. (km)"]
FLATFILE_COLS['HypD'] = ["HypD (km)"]
FLATFILE_COLS['ClstD'] = ["ClstD (km)"]
FLATFILE_COLS['PGA'] = ["PGA (g)"]
FLATFILE_COLS['PGV'] = ["PGV (cm/sec)"]
FLATFILE_COLS['Tp'] = ["Tp"]

# for ML:
FEATURE_KEYS = [k for k in FLATFILE_COLS.keys() if k != "RSN"]  # exclude RSN  # edit as you like
USE_INTERCEPT = True
LOG_KEYS = set(FEATURE_KEYS)  # e.g., {"VS30", "Rjb"} if you want log transforms

def load_flatfile_xlsx_rows(path: str, sheet: Optional[str] = None) -> List[Dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.worksheets[0]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        raise ValueError(f"XLSX has no rows: {path}")

    headers = [str(h).strip() if h is not None else "" for h in header]
    if all(h == "" for h in headers):
        raise ValueError("Header row appears empty in XLSX.")

    out: List[Dict[str, str]] = []
    for r in rows_iter:
        d: Dict[str, str] = {}
        for k, v in zip(headers, r):
            if k == "":
                continue
            d[k] = "" if v is None else str(v).strip()
        if any(val != "" for val in d.values()):
            out.append(d)

    if len(out) == 0:
        raise ValueError(f"No data rows found in XLSX: {path}")
    return out



def build_flat_metadata_map(flat_rows: List[Dict[str, str]]) -> Dict[str, Dict[str, float]]:
    """
    Map rsn_int_str -> flat_vars dict (keys from FLATFILE_COLS excluding RSN).
    dt is NOT included (comes from HDF5).
    """
    meta: Dict[str, Dict[str, float]] = {}

    for r in flat_rows:
        rsn = _pick_col_value(r, FLATFILE_COLS["RSN"], required=False)
        if rsn is None:
            continue
        try:
            rsn_key = str(int(float(rsn)))
        except Exception:
            continue

        flat_vars: Dict[str, float] = {}
        for key, candidates in FLATFILE_COLS.items():
            if key == "RSN":
                continue
            flat_vars[key] = _to_float_nan999(_pick_col_value(r, candidates, required=False))


        meta[rsn_key] = flat_vars

    if len(meta) == 0:
        raise ValueError("No RSNs parsed from flatfile XLSX. Check RSN columns.")
    return meta



# ---------------------------
# Filename-mapping CSV (H1/H2 only)
# ---------------------------

# Allow some variation in RSN column naming
FILENAME_TABLE_RSN_COLS = ["NGARSN", "RSN", "Record Sequence Number", "RecordSequenceNumber"]

# You said mapping csv has NGAFilename_H1 / NGAFilename_H2
MAP_H1_COLS = ["NGAFilename_H1", "Filename_H1", "H1", "FileH1"]
MAP_H2_COLS = ["NGAFilename_H2", "Filename_H2", "H2", "FileH2"]


def load_rsn_to_h1h2_dataset_names(filename_csv: str) -> Dict[str, Tuple[str, str]]:
    """
    Return rsn -> (nameH1, nameH2) using ONLY mapping CSV.

    The values may be dataset names or stems (may omit '.AT2').
    We keep raw strings and handle matching against HDF5 later.
    """
    rows = load_csv_rows(filename_csv)
    out: Dict[str, Tuple[str, str]] = {}

    for r in rows:
        rsn = _pick_col_value(r, FILENAME_TABLE_RSN_COLS, required=False)
        if rsn is None:
            continue
        try:
            rsn_key = str(int(float(rsn)))
        except Exception:
            continue

        h1 = _pick_col_value(r, MAP_H1_COLS, required=False)
        h2 = _pick_col_value(r, MAP_H2_COLS, required=False)

        # Require BOTH to be valid here (per your rule)
        if h1 is None or h2 is None:
            continue
        if _is_missing_token(h1) or _is_missing_token(h2):
            continue

        out[rsn_key] = (h1.strip(), h2.strip())

    if len(out) == 0:
        raise ValueError(
            f"No valid H1/H2 mappings found in {filename_csv}. "
            f"Expected columns like NGARSN, NGAFilename_H1, NGAFilename_H2."
        )
    return out


# ---------------------------
# HDF5 lookup
# ---------------------------

def _candidate_dataset_names(mapped: str) -> List[str]:
    """
    For this HDF5, the dataset names are typically mapping_value + '.AT2'.

    Example mapping_value: RSN19923_40190509_1830HNN
    Example dataset name:  RSN19923_40190509_1830HNN.AT2
    """
    s = mapped.strip().strip('"').strip("'")
    s = s.replace("\\", "_").replace("/", "_")

    cands = []
    # Prefer .AT2 first (your HDF5 uses it)
    if not s.lower().endswith(".at2"):
        cands.append(s + ".AT2")
        cands.append(s + ".at2")
    # Then try raw (in case some datasets were stored without extension)
    cands.append(s)

    # de-dup preserve order
    out = []
    seen = set()
    for x in cands:
        if x not in seen:
            out.append(x)
            seen.add(x)
    return out




def _find_dataset_path(h5: h5py.File, rsn_int_str: str, mapped_name: str) -> Optional[str]:
    """
    Looks for dataset under /RSN{rsn}/<candidate_name>.
    Returns full path if found else None.
    """
    group = f"/RSN{rsn_int_str}"
    if group not in h5:
        return None

    for dsname in _candidate_dataset_names(mapped_name):
        dspath = f"{group}/{dsname}"
        if dspath in h5:
            return dspath
    return None


# ---------------------------
# Metrics + logging
# ---------------------------



@dataclass
class RecordMetrics:
    rec_id: str
    component: str   # H1 or H2
    h5_ref: str

    # flatfile variables (whatever keys you define in FLATFILE_COLS, except RSN)
    flat: Dict[str, float] = field(default_factory=dict)

    # waveform-derived (dt comes from HDF5 attr)
    dt: float = np.nan
    npts: int = 0
    duration: float = np.nan
    amax: float = np.nan
    amin: float = np.nan
    amp_range: float = np.nan
    tmax: float = np.nan
    tmin: float = np.nan
    dt_peaks: float = np.nan
    dt_peaks_norm: float = np.nan

    ok: int = 1
    err: str = ""



@dataclass
class RecordLog:
    rsn: str
    component: str
    mapped_name: str
    h5_dataset: str
    status: str   # processed | skipped
    reason: str


def _metrics_from_array(
    rsn: str,
    component: str,
    h5_ref: str,
    a: np.ndarray,
    dt: float,
    flat_vars: Dict[str, float],
) -> RecordMetrics:
    npts = int(a.size)
    amax = float(np.max(a))
    amin = float(np.min(a))
    amp_range = float(amax - amin)

    if np.isfinite(dt) and dt > 0:
        imax = int(np.argmax(a))
        imin = int(np.argmin(a))
        tmax = float(imax * dt)
        tmin = float(imin * dt)
        dt_peaks = float(abs(tmax - tmin))
        duration = float((npts - 1) * dt) if npts > 1 else 0.0
        dt_peaks_norm = float(dt_peaks / duration) if duration > 0 else np.nan
    else:
        tmax = tmin = dt_peaks = duration = dt_peaks_norm = np.nan

    return RecordMetrics(
        rec_id=rsn,
        component=component,
        h5_ref=h5_ref,
        flat=dict(flat_vars),   # snapshot
        dt=dt,
        npts=npts,
        duration=duration,
        amax=amax,
        amin=amin,
        amp_range=amp_range,
        tmax=tmax,
        tmin=tmin,
        dt_peaks=dt_peaks,
        dt_peaks_norm=dt_peaks_norm,
        ok=1,
        err="",
    )



# ---------------------------
# ML helpers + output
# ---------------------------
# Define your ML features here, once.
# These keys must match the keys in your FLATFILE_COLS dict (except RSN).

# Option A (minimal): keep dropping rows for ML, but only ML
# This is simplest and totally consistent with your “don’t skip records” goal — because you’re not skipping the record, only deciding it can’t be used for training.
# If you choose this option, just keep make_features() as-is (after the -999 -> NaN change above), and maybe add a count so you see how many training rows survived.
# Option B (tolerate missing predictors): mean/median impute + missing indicators ("impute")
# If you want to keep training rows even with missing flatfile fields, do this:
# for each feature k, add:
# k value with imputation (median is fine)
# is_missing_k indicator (0/1)
# That lets the regression learn that “missing Tp” is a different regime.

from typing import NamedTuple


from typing import NamedTuple, Any
import json

class MLBuildInfo(NamedTuple):
    candidates: int
    used: int
    dropped_missing: int
    dropped_outlier: int


def _max_abs_amp(m: RecordMetrics) -> float:
    # robust abs-max based on already-computed extrema
    if np.isfinite(m.amax) and np.isfinite(m.amin):
        return float(max(abs(m.amax), abs(m.amin)))
    return np.nan


def make_features_with_ids(
    metrics: List[RecordMetrics],
    *,
    mode: str = "drop",
    max_abs_amp: float = 10.0,
) -> Tuple[np.ndarray, np.ndarray, List[str], List[str], MLBuildInfo]:
    """
    Build X,y for predicting amp_range from flatfile vars.
    Returns:
      X, y, feature_names, ids_used, info
    Screening:
      - ML skips outliers where max(|a|) > max_abs_amp (but metrics are still written).
    """
    if mode not in {"drop", "impute"}:
        raise ValueError(f"Unknown mode: {mode}")

    # 1) candidate rows must have finite target AND pass outlier screen
    rows_raw: List[Dict[str, float]] = []
    y: List[float] = []
    ids: List[str] = []

    dropped_outlier = 0

    for m in metrics:
        if m.ok != 1 or not np.isfinite(m.amp_range):
            continue

        aa = _max_abs_amp(m)
        if np.isfinite(aa) and aa > float(max_abs_amp):
            dropped_outlier += 1
            continue

        rows_raw.append(dict(m.flat or {}))
        y.append(float(m.amp_range))
        ids.append(str(m.rec_id))

    candidates = len(y)
    if candidates == 0:
        feature_names: List[str] = []
        if USE_INTERCEPT:
            feature_names.append("bias")
        feature_names.extend(FEATURE_KEYS)
        return (
            np.zeros((0, len(feature_names))), np.zeros((0,)),
            feature_names, [], MLBuildInfo(0, 0, 0, dropped_outlier)
        )

    yv = np.asarray(y, dtype=float)

    # 2) build raw feature matrix in model-space (log keys -> ln)
    F = len(FEATURE_KEYS)
    X0 = np.full((candidates, F), np.nan, dtype=float)

    for i, fv in enumerate(rows_raw):
        for j, k in enumerate(FEATURE_KEYS):
            v = fv.get(k, np.nan)
            v = float(v) if v is not None else np.nan
            if k in LOG_KEYS:
                if np.isfinite(v) and v > 0:
                    v = math.log(v)
                else:
                    v = np.nan
            X0[i, j] = v

    # 3) missing handling
    if mode == "drop":
        keep = np.isfinite(X0).all(axis=1) & np.isfinite(yv)
        Xkeep = X0[keep]
        ykeep = yv[keep]
        ids_keep = [ids[i] for i in range(candidates) if bool(keep[i])]

        if USE_INTERCEPT:
            X = np.column_stack([np.ones((Xkeep.shape[0], 1)), Xkeep])
            feature_names = ["bias"] + list(FEATURE_KEYS)
        else:
            X = Xkeep
            feature_names = list(FEATURE_KEYS)

        dropped_missing = int(candidates - Xkeep.shape[0])
        info = MLBuildInfo(candidates=candidates, used=int(X.shape[0]),
                           dropped_missing=dropped_missing, dropped_outlier=dropped_outlier)
        return X, ykeep, feature_names, ids_keep, info

    # mode == "impute"
    miss = ~np.isfinite(X0)

    Ximp = X0.copy()
    for j in range(F):
        col = Ximp[:, j]
        good = np.isfinite(col)
        med = float(np.nanmedian(col[good])) if np.any(good) else 0.0
        col[~good] = med
        Ximp[:, j] = col

    parts = []
    feature_names = []

    if USE_INTERCEPT:
        parts.append(np.ones((candidates, 1), dtype=float))
        feature_names.append("bias")

    parts.append(Ximp)
    feature_names.extend(list(FEATURE_KEYS))

    M = miss.astype(float)
    parts.append(M)
    feature_names.extend([f"miss_{k}" for k in FEATURE_KEYS])

    X = np.column_stack(parts)
    keep_y = np.isfinite(yv)
    X = X[keep_y]
    yv2 = yv[keep_y]
    ids_keep = [ids[i] for i in range(candidates) if bool(keep_y[i])]

    dropped_missing = int(candidates - X.shape[0])  # usually 0 here unless y had NaNs
    info = MLBuildInfo(candidates=candidates, used=int(X.shape[0]),
                       dropped_missing=dropped_missing, dropped_outlier=dropped_outlier)
    return X, yv2, feature_names, ids_keep, info


def _train_test_split_mask(n: int, test_frac: float, seed: int) -> np.ndarray:
    rng = np.random.default_rng(int(seed))
    idx = np.arange(n)
    rng.shuffle(idx)
    ntest = int(np.floor(float(test_frac) * n))
    test = np.zeros(n, dtype=bool)
    if ntest > 0:
        test[idx[:ntest]] = True
    return test


def fit_linear_model(Xtr: np.ndarray, ytr: np.ndarray) -> Tuple[Any, np.ndarray]:
    """
    Fit and return (model_or_coef, coef_vector).
    Prefer sklearn LinearRegression(fit_intercept=False), fallback to lstsq.
    """
    try:
        from sklearn.linear_model import LinearRegression  # type: ignore
        model = LinearRegression(fit_intercept=False)
        model.fit(Xtr, ytr)
        coef = np.asarray(model.coef_, dtype=float)
        return model, coef
    except Exception:
        coef, *_ = np.linalg.lstsq(Xtr, ytr, rcond=None)
        return None, np.asarray(coef, dtype=float)


def r2_score(y: np.ndarray, yhat: np.ndarray) -> float:
    y = np.asarray(y, dtype=float)
    yhat = np.asarray(yhat, dtype=float)
    m = np.isfinite(y) & np.isfinite(yhat)
    if np.sum(m) < 3:
        return np.nan
    y = y[m]; yhat = yhat[m]
    ss_res = float(np.sum((y - yhat) ** 2))
    ss_tot = float(np.sum((y - np.mean(y)) ** 2))
    return 1.0 - ss_res / ss_tot if ss_tot > 0 else np.nan


def save_model_json(path: str, *, feature_names: List[str], coef: np.ndarray, meta: Dict[str, Any]) -> None:
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    blob = {
        "type": "linear_regression",
        "feature_names": list(feature_names),
        "coef": [float(c) for c in np.asarray(coef, dtype=float).ravel().tolist()],
        "meta": meta,
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(blob, f, indent=2)


def save_model_joblib(path: str, model: Any, *, meta: Dict[str, Any]) -> None:
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    try:
        import joblib  # type: ignore
        payload = {"model": model, "meta": meta}
        joblib.dump(payload, path)
    except Exception as e:
        raise RuntimeError(f"joblib save failed: {e}")


def make_features(
    metrics: List[RecordMetrics],
    *,
    mode: str = "drop",
) -> Tuple[np.ndarray, np.ndarray, List[str], MLBuildInfo]:
    """
    Build X,y for predicting amp_range from flatfile variables.

    mode:
      - "drop": drop any row with any missing feature after transforms
      - "impute": median-impute missing features + add missing-indicator columns

    Features are defined by FEATURE_KEYS (and optional intercept USE_INTERCEPT).
    Optional log-transform controlled by LOG_KEYS.

    Returns: (X, y, feature_names, info)
    """
    if mode not in {"drop", "impute"}:
        raise ValueError(f"Unknown mode: {mode}")

    # 1) collect candidate rows (must have finite target)
    rows_raw: List[Dict[str, float]] = []
    y: List[float] = []

    for m in metrics:
        if m.ok != 1 or not np.isfinite(m.amp_range):
            continue
        rows_raw.append(dict(m.flat or {}))
        y.append(float(m.amp_range))

    candidates = len(y)
    if candidates == 0:
        feature_names: List[str] = []
        if USE_INTERCEPT:
            feature_names.append("bias")
        feature_names.extend(FEATURE_KEYS)
        return np.zeros((0, len(feature_names))), np.zeros((0,)), feature_names, MLBuildInfo(0, 0, 0)

    yv = np.asarray(y, dtype=float)

    # 2) build raw feature matrix (float), applying log transform -> may create NaNs
    #    We do transforms first, then missing handling.
    F = len(FEATURE_KEYS)
    X0 = np.full((candidates, F), np.nan, dtype=float)

    for i, fv in enumerate(rows_raw):
        for j, k in enumerate(FEATURE_KEYS):
            v = fv.get(k, np.nan)
            v = float(v) if v is not None else np.nan
            if k in LOG_KEYS:
                if np.isfinite(v) and v > 0:
                    v = math.log(v)
                else:
                    v = np.nan
            X0[i, j] = v

    # 3) missing handling
    if mode == "drop":
        keep = np.isfinite(X0).all(axis=1) & np.isfinite(yv)
        Xkeep = X0[keep]
        ykeep = yv[keep]

        # add intercept if needed
        if USE_INTERCEPT:
            X = np.column_stack([np.ones((Xkeep.shape[0], 1)), Xkeep])
            feature_names = ["bias"] + list(FEATURE_KEYS)
        else:
            X = Xkeep
            feature_names = list(FEATURE_KEYS)

        info = MLBuildInfo(candidates=candidates, used=int(X.shape[0]), dropped=int(candidates - X.shape[0]))
        return X, ykeep, feature_names, info

    # mode == "impute"
    # Missing indicator computed on transformed space (after log rules).
    miss = ~np.isfinite(X0)

    # Median impute (per feature) using available values
    Ximp = X0.copy()
    for j in range(F):
        col = Ximp[:, j]
        good = np.isfinite(col)
        if np.any(good):
            med = float(np.nanmedian(col[good]))
        else:
            # feature entirely missing across all candidates -> impute 0.0
            # (indicator will carry the info)
            med = 0.0
        col[~good] = med
        Ximp[:, j] = col

    # Build final X:
    # [bias?] + features + missing_indicators
    parts = []
    feature_names = []

    if USE_INTERCEPT:
        parts.append(np.ones((candidates, 1), dtype=float))
        feature_names.append("bias")

    parts.append(Ximp)
    feature_names.extend(list(FEATURE_KEYS))

    # Add missing indicators (0/1) for each feature key
    M = miss.astype(float)
    parts.append(M)
    feature_names.extend([f"miss_{k}" for k in FEATURE_KEYS])

    X = np.column_stack(parts)
    keep = np.isfinite(yv)  # should already be true; keep for safety
    X = X[keep]
    yv2 = yv[keep]

    info = MLBuildInfo(candidates=candidates, used=int(X.shape[0]), dropped=int(candidates - X.shape[0]))
    return X, yv2, feature_names, info





def fit_and_report_regression(X: np.ndarray, y: np.ndarray, feature_names: List[str]) -> Dict[str, float]:
    report: Dict[str, float] = {}
    if X.shape[0] < X.shape[1] + 2:
        report["r2"] = np.nan
        return report
    try:
        from sklearn.linear_model import LinearRegression  # type: ignore
        model = LinearRegression(fit_intercept=False)
        model.fit(X, y)
        yhat = model.predict(X)
        ss_res = float(np.sum((y - yhat) ** 2))
        ss_tot = float(np.sum((y - np.mean(y)) ** 2))
        report["r2"] = 1.0 - ss_res / ss_tot if ss_tot > 0 else np.nan
        for name, coef in zip(feature_names, model.coef_.tolist()):
            report[f"coef_{name}"] = float(coef)
        return report
    except Exception:
        coef, *_ = np.linalg.lstsq(X, y, rcond=None)
        yhat = X @ coef
        ss_res = float(np.sum((y - yhat) ** 2))
        ss_tot = float(np.sum((y - np.mean(y)) ** 2))
        report["r2"] = 1.0 - ss_res / ss_tot if ss_tot > 0 else np.nan
        for name, c in zip(feature_names, coef.tolist()):
            report[f"coef_{name}"] = float(c)
        return report




from dataclasses import fields, is_dataclass

def write_preds_csv(path: str, rows: List[Dict[str, Any]]) -> None:
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    if not rows:
        # still write header
        with open(path, "w", newline="", encoding="utf-8") as f:
            f.write("rec_id,split,y,yhat,resid\n")
        return

    # stable field order
    fieldnames = list(rows[0].keys())
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in rows:
            w.writerow(r)


def write_metrics_csv(path: str, metrics: list[RecordMetrics]) -> None:
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)

    # Base dataclass fields
    dfields = [f.name for f in fields(RecordMetrics)]

    # Auto-discover keys in any dict-valued fields (like "flat")
    dict_keys: dict[str, list[str]] = {}
    for m in metrics:
        for name in dfields:
            v = getattr(m, name)
            if isinstance(v, dict):
                s = dict_keys.setdefault(name, [])
                for k in v.keys():
                    if k not in s:
                        s.append(k)

    # Final CSV header:
    # - include all non-dict fields as-is
    # - expand dict fields into <field>__<key>
    fieldnames: list[str] = []
    for name in dfields:
        if name in dict_keys:
            for k in dict_keys[name]:
                fieldnames.append(f"{name}__{k}")
        else:
            fieldnames.append(name)

    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()

        for m in metrics:
            row = {}
            for name in dfields:
                v = getattr(m, name)
                if name in dict_keys:
                    vv = v if isinstance(v, dict) else {}
                    for k in dict_keys[name]:
                        row[f"{name}__{k}"] = vv.get(k, np.nan)
                else:
                    row[name] = v
            w.writerow(row)




def write_logs_csv(path: str, logs: List[RecordLog]) -> None:
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    fieldnames = list(RecordLog.__annotations__.keys())
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for l in logs:
            w.writerow({k: getattr(l, k) for k in fieldnames})


def write_report_txt(
    path: str,
    attempted_units: int,
    processed_units: int,
    skipped_units: int,
    Xrows: int,
    report: Dict[str, float],
    *,
    unit_label: str = "component-records",
) -> None:
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        f.write(f"Attempted ({unit_label}): {attempted_units}\n")
        f.write(f"Processed ({unit_label}): {processed_units}\n")
        f.write(f"Skipped   ({unit_label}): {skipped_units}\n")
        f.write(f"Training rows (after NaN filter): {Xrows}\n\n")
        
        f.write(f"R^2 train: {report.get('r2_train', np.nan):.3g}\n")
        f.write(f"R^2 test : {report.get('r2_test', np.nan):.3g}\n\n")
        f.write(f"ML candidates (after target+ok+outlier screen): {report.get('n_candidates', 0)}\n")
        f.write(f"ML used rows: {report.get('n_used', 0)}\n")
        f.write(f"Dropped (missing): {report.get('dropped_missing', 0)}\n")
        f.write(f"Dropped (outlier max_abs_amp): {report.get('dropped_outlier', 0)}\n")
        f.write(f"Mode: {report.get('mode', '')}\n")
        f.write(f"max_abs_amp: {report.get('max_abs_amp', np.nan)}\n")
        f.write(f"test_frac: {report.get('test_frac', np.nan)}\n")
        f.write(f"seed: {report.get('seed', '')}\n")
       
        for k, v in report.items():
            if k.startswith("coef_"):
                f.write(f"{k}: {v:.3g}\n")



# ---------------------------
# Main (MPI)
# ---------------------------

def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--flatfile-xlsx", required=True, help="Flatfile .xlsx (filenames ignored).")
    ap.add_argument("--flatfile-sheet", default="", help="Optional sheet name (default: first sheet).")
    ap.add_argument("--filenames-csv", required=True, help="Filename-mapping CSV (use NGAFilename_H1/H2).")
    ap.add_argument("--hdf5", required=True, help="Merged TimeSeries HDF5 file")

    ap.add_argument("--out-prefix", default="out_", help="Prefix for output files")
    ap.add_argument("--limit", type=int, default=-1, help="Optional RSN limit for debugging")

    ap.add_argument("--compute-rotd", action="store_true", help="Compute RotD0/50/100 using H1+H2 rotations")
    ap.add_argument("--rotd-angle-step", type=float, default=1.0, help="Angle step in degrees (default 1 deg)")

    ap.add_argument("--outdir", default=".", help="Directory to write all outputs (default: current dir)")

    ap.add_argument(
        "--ml-missing",
        choices=["drop", "impute"],
        default="drop",
        help="How to handle missing flatfile predictors for ML: "
             "'drop' drops rows with any missing feature; "
             "'impute' keeps rows by median-imputing and adding missing indicators."
    )

    ap.add_argument("--test-frac", type=float, default=0.20,
                    help="Test fraction for ML split (default 0.20).")
    ap.add_argument("--seed", type=int, default=12345,
                    help="RNG seed for ML split.")

    ap.add_argument("--max-abs-amp", type=float, default=10.0,
                    help="Outlier screen for ML only: skip records where max(|a|) > this (default 10).")

    ap.add_argument("--write-preds", action="store_true",
                    help="Write per-record predictions/residuals CSV for ML rows.")

    ap.add_argument("--save-model", choices=["none", "json", "joblib"], default="json",
                    help="Serialize fitted ML model. json saves coefficients+metadata; joblib saves sklearn model (if available).")

    ap.add_argument("--model-path", default="",
                    help="Optional output path for model artifact. If empty, uses <outdir>/<prefix>_model_<comp>.(json|joblib).")

    
    args = ap.parse_args()

    angles_deg = np.arange(0.0, 180.0, args.rotd_angle_step, dtype=float)

    # print('args.flatfile_xlsx',args.flatfile_xlsx)
    # print('args.filenames_csv',args.filenames_csv)
    # print('args.hdf5',args.hdf5)
    _ensure_file_exists(args.flatfile_xlsx, "Flatfile XLSX")
    _ensure_file_exists(args.filenames_csv, "Filename-mapping CSV")
    _ensure_file_exists(args.hdf5, "HDF5 file")

    comm = MPI.COMM_WORLD
    rank = comm.Get_rank()
    size = comm.Get_size()


    outdir = None
    if rank == 0:
        outdir = Path(args.outdir).expanduser().resolve()
        outdir.mkdir(parents=True, exist_ok=True)
    outdir = comm.bcast(outdir, root=0)
    

    if rank == 0:
        flat_rows = load_flatfile_xlsx_rows(args.flatfile_xlsx, sheet=(args.flatfile_sheet or None))
        flat_meta = build_flat_metadata_map(flat_rows)  # rsn -> dict of flatfile variables (no DT)


        rsn_to_h1h2 = load_rsn_to_h1h2_dataset_names(args.filenames_csv)

        # Attempt RSNs from flatfile (so we have Mw/Vs30/Dist)
        rsn_list = sorted(flat_meta.keys(), key=lambda s: int(s))
        if args.limit > 0:
            rsn_list = rsn_list[: args.limit]

        n_total = len(rsn_list)
    else:
        flat_meta = None
        rsn_to_h1h2 = None
        rsn_list = None
        n_total = None

    n_total = comm.bcast(n_total, root=0)
    flat_meta = comm.bcast(flat_meta, root=0)
    rsn_to_h1h2 = comm.bcast(rsn_to_h1h2, root=0)
    rsn_list = comm.bcast(rsn_list, root=0)

    if rank == 0:
        indices = np.arange(n_total, dtype=int)
        chunks = np.array_split(indices, size)
        chunks_list = [c.tolist() for c in chunks]
    else:
        chunks_list = None

    my_idx: List[int] = comm.scatter(chunks_list, root=0)

    # We output separate metrics/logs for H1 and H2
    my_metrics = {"H1": [], "H2": []}   # type: ignore[var-annotated]
    my_logs = {"H1": [], "H2": []}      # type: ignore[var-annotated]

    rotd_rows_local: List[RotDRow] = []


    with h5py.File(args.hdf5, "r") as h5:
        for ii in my_idx:
            rsn = rsn_list[ii]
            flat_vars = flat_meta[rsn]  # dict of variables from flatfile

            # Rule: only skip RSN if H1 or H2 are missing (or missing in mapping)
            pair = (rsn_to_h1h2 or {}).get(rsn, None)  # type: ignore[union-attr]
            if pair is None:
                # Log skip for both
                for comp in ("H1", "H2"):
                    my_logs[comp].append(
                        RecordLog(rsn=rsn, component=comp, mapped_name="", h5_dataset="",
                                  status="skipped", reason="missing_h1h2_mapping_for_rsn")
                    )
                continue

            mapped_h1, mapped_h2 = pair

            # Find datasets for BOTH. If either missing, skip RSN (both comps)
            dspath_h1 = _find_dataset_path(h5, rsn, mapped_h1)
            dspath_h2 = _find_dataset_path(h5, rsn, mapped_h2)

            if dspath_h1 is None or dspath_h2 is None:
                # Log separately with precise reason
                if dspath_h1 is None:
                    my_logs["H1"].append(
                        RecordLog(rsn=rsn, component="H1", mapped_name=mapped_h1, h5_dataset="",
                                  status="skipped", reason="h1_dataset_not_found_in_hdf5")
                    )
                else:
                    my_logs["H1"].append(
                        RecordLog(rsn=rsn, component="H1", mapped_name=mapped_h1, h5_dataset=dspath_h1,
                                  status="skipped", reason="skipped_because_h2_missing")
                    )

                if dspath_h2 is None:
                    my_logs["H2"].append(
                        RecordLog(rsn=rsn, component="H2", mapped_name=mapped_h2, h5_dataset="",
                                  status="skipped", reason="h2_dataset_not_found_in_hdf5")
                    )
                else:
                    my_logs["H2"].append(
                        RecordLog(rsn=rsn, component="H2", mapped_name=mapped_h2, h5_dataset=dspath_h2,
                                  status="skipped", reason="skipped_because_h1_missing")
                    )
                continue

            # If we get here: BOTH datasets exist -> process both
            
            ds1 = h5[dspath_h1]
            ds2 = h5[dspath_h2]
            
            # Prefer dtHeader from H1; fallback to H2; fallback to flatfile dt
            dt1 = _get_dt_from_dataset(ds1)
            dt2 = _get_dt_from_dataset(ds2)
            
            dt_use = dt1
            if not (np.isfinite(dt_use) and dt_use > 0):
                dt_use = dt2
            # no flatfile DT fallback
            if not (np.isfinite(dt_use) and dt_use > 0):
                dt_use = np.nan


            
            a1 = a2 = None
            ok_a1 = ok_a2 = False
            
            # H1
            try:
                a1 = np.asarray(ds1, dtype=float).ravel()
                ok_a1 = True
                h5_ref1 = f"{args.hdf5}:{dspath_h1}"
                m1 = _metrics_from_array(rsn, "H1", h5_ref1, a1, dt_use, flat_vars)
                my_metrics["H1"].append(m1)
                my_logs["H1"].append(
                    RecordLog(rsn=rsn, component="H1", mapped_name=mapped_h1, h5_dataset=dspath_h1,
                              status="processed", reason="")
                )
            except Exception as e:
                my_logs["H1"].append(
                    RecordLog(rsn=rsn, component="H1", mapped_name=mapped_h1, h5_dataset=dspath_h1,
                              status="skipped", reason=f"h1_read_or_compute_error:{e}")
                )
            
            # H2
            try:
                a2 = np.asarray(ds2, dtype=float).ravel()
                ok_a2 = True
                h5_ref2 = f"{args.hdf5}:{dspath_h2}"
                m2 = _metrics_from_array(rsn, "H2", h5_ref2, a2, dt_use, flat_vars)
                my_metrics["H2"].append(m2)
                my_logs["H2"].append(
                    RecordLog(rsn=rsn, component="H2", mapped_name=mapped_h2, h5_dataset=dspath_h2,
                              status="processed", reason="")
                )
            except Exception as e:
                my_logs["H2"].append(
                    RecordLog(rsn=rsn, component="H2", mapped_name=mapped_h2, h5_dataset=dspath_h2,
                              status="skipped", reason=f"h2_read_or_compute_error:{e}")
                )
            
            # RotD (only if BOTH arrays exist)
            if args.compute_rotd and ok_a1 and ok_a2:
                rr = compute_rotd_summaries(a1, a2, dt_use, angles_deg)
                for row in rr:
                    row.rsn = rsn
                    rotd_rows_local.append(row)




    # Gather
    all_metrics_h1 = comm.gather(my_metrics["H1"], root=0)
    all_metrics_h2 = comm.gather(my_metrics["H2"], root=0)
    all_logs_h1 = comm.gather(my_logs["H1"], root=0)
    all_logs_h2 = comm.gather(my_logs["H2"], root=0)
    all_rotd = comm.gather(rotd_rows_local, root=0)


    if rank == 0:
        for comp, gathered_metrics, gathered_logs in [
            ("H1", all_metrics_h1, all_logs_h1),
            ("H2", all_metrics_h2, all_logs_h2),
        ]:
            metrics = [m for sub in gathered_metrics for m in sub]
            logs = [l for sub in gathered_logs for l in sub]


            # rank 0
            outdir = resolve_outdir(args.outdir)
            
            out_metrics   = outpath(outdir, args.out_prefix, f"metrics_{comp}", ".csv")
            out_processed = outpath(outdir, args.out_prefix, f"processed_{comp}", ".csv")
            out_skipped   = outpath(outdir, args.out_prefix, f"skipped_{comp}", ".csv")
            out_report    = outpath(outdir, args.out_prefix, f"ml_report_{comp}", ".txt")
            
            


            write_metrics_csv(str(out_metrics), metrics)
            write_logs_csv(str(out_processed), [l for l in logs if l.status == "processed"])
            write_logs_csv(str(out_skipped), [l for l in logs if l.status == "skipped"])
            


            # train/test + save-model + write-preds block
            # --- ML build (screen outliers ONLY for ML) ---
            X, y, feat_names, ids_used, info = make_features_with_ids(
                metrics,
                mode=args.ml_missing,
                max_abs_amp=float(args.max_abs_amp),
            )

            # ML split
            test_mask = _train_test_split_mask(X.shape[0], test_frac=float(args.test_frac), seed=int(args.seed))
            train_mask = ~test_mask

            report = {
                "r2_train": np.nan,
                "r2_test": np.nan,
                "n_candidates": int(info.candidates),
                "n_used": int(info.used),
                "dropped_missing": int(info.dropped_missing),
                "dropped_outlier": int(info.dropped_outlier),
                "mode": str(args.ml_missing),
                "max_abs_amp": float(args.max_abs_amp),
                "test_frac": float(args.test_frac),
                "seed": int(args.seed),
            }

            model_obj = None
            coef = None

            if X.shape[0] >= (X.shape[1] + 2) and np.sum(train_mask) >= (X.shape[1] + 2):
                Xtr, ytr = X[train_mask], y[train_mask]
                Xte, yte = X[test_mask], y[test_mask]

                model_obj, coef = fit_linear_model(Xtr, ytr)
                yhat_tr = Xtr @ coef
                yhat_te = Xte @ coef if Xte.size else np.array([], dtype=float)

                report["r2_train"] = float(r2_score(ytr, yhat_tr))
                report["r2_test"] = float(r2_score(yte, yhat_te)) if Xte.size else np.nan
            else:
                # not enough rows to fit
                coef = np.full((X.shape[1],), np.nan, dtype=float)

            # --- optional: write predictions ---
            if args.write_preds and X.shape[0] > 0 and coef is not None and np.all(np.isfinite(coef)):
                yhat_all = X @ coef
                rows_out: List[Dict[str, Any]] = []
                for i in range(X.shape[0]):
                    split = "test" if bool(test_mask[i]) else "train"
                    rows_out.append({
                        "rec_id": ids_used[i],
                        "component": comp,
                        "split": split,
                        "y": float(y[i]),
                        "yhat": float(yhat_all[i]),
                        "resid": float(y[i] - yhat_all[i]),
                    })
                out_preds = outpath(outdir, args.out_prefix, f"ml_preds_{comp}", ".csv")
                write_preds_csv(str(out_preds), rows_out)
                print(f"[rank 0] wrote: {out_preds}")

            # --- optional: serialize model ---
            if args.save_model != "none" and coef is not None and np.all(np.isfinite(coef)):
                meta = {
                    "prefix": str(args.out_prefix),
                    "component": str(comp),
                    "ml_missing": str(args.ml_missing),
                    "feature_keys": list(FEATURE_KEYS),
                    "log_keys": sorted(list(LOG_KEYS)),
                    "use_intercept": bool(USE_INTERCEPT),
                    "max_abs_amp": float(args.max_abs_amp),
                    "test_frac": float(args.test_frac),
                    "seed": int(args.seed),
                    "report": report,
                }

                if args.model_path.strip():
                    model_path = str(Path(args.model_path).expanduser().resolve())
                else:
                    ext = ".json" if args.save_model == "json" else ".joblib"
                    model_path = str(outpath(outdir, args.out_prefix, f"model_{comp}", ext))

                if args.save_model == "json":
                    save_model_json(model_path, feature_names=feat_names, coef=coef, meta=meta)
                    print(f"[rank 0] wrote model: {model_path}")
                elif args.save_model == "joblib":
                    if model_obj is None:
                        raise RuntimeError("Requested joblib model save, but sklearn model object not available (fallback lstsq used). Use --save-model json.")
                    save_model_joblib(model_path, model_obj, meta=meta)
                    print(f"[rank 0] wrote model: {model_path}")




            


            
            # Optional: print ML row accounting
            
            print(
                f"[rank 0] {comp} ML rows: candidates={info.candidates} used={info.used} "
                f"dropped_missing={info.dropped_missing} dropped_outlier={info.dropped_outlier} "
                f"mode={args.ml_missing}"
            )



            attempted = len(logs)
            processed = len([l for l in logs if l.status == "processed"])
            skipped = attempted - processed
            
            attempted_rsn = len(set(l.rsn for l in logs))
            processed_rsn = len(set(l.rsn for l in logs if l.status == "processed"))
            skipped_rsn = attempted_rsn - processed_rsn

            # write_report_txt(out_report, attempted, processed, skipped, X.shape[0], report)
            write_report_txt(str(out_report), attempted, processed, skipped, X.shape[0], report, unit_label="component-rows")

            print(
                f"[rank 0] {comp}: attempted={attempted} processed={processed} skipped={skipped} "
                f"train_rows={X.shape[0]} R2_train={report.get('r2_train', np.nan):.3g} "
                f"R2_test={report.get('r2_test', np.nan):.3g}"
            )

            # print(
            #     f"[rank 0] {comp}: attempted={attempted} processed={processed} "
            #     f"skipped={skipped} train_rows={X.shape[0]} R2={report.get('r2', np.nan):.3g}"
            # )
            print(f"[rank 0] wrote: {out_metrics}")
            print(f"[rank 0] wrote: {out_processed}")
            print(f"[rank 0] wrote: {out_skipped}")
            print(f"[rank 0] wrote: {out_report}")

        if args.compute_rotd:
            rotd_rows = [r for sub in all_rotd for r in sub]
            out_rotd = outpath(outdir, args.out_prefix, "metrics_RotD", ".csv")
            write_rotd_csv(str(out_rotd), rotd_rows)
            print(f"[rank 0] wrote: {out_rotd}")



    return 0


if __name__ == "__main__":
    raise SystemExit(main())


# NOTE: This is a time-domain RotD analog (based on max-abs acceleration and other time-domain scalars). It’s conceptually aligned with RotD, but it’s not the same as NGA’s RotD of response spectra unless you compute PSA for each rotation/period.

# ### How to run

# mpirun -np 8 python3 nga_mpi_metrics_ml_hdf5.py \
#   --flatfile-xlsx /home/jupyter/Work/stampede3/Datasets/NGAWest2/Updated_NGA_West2_Flatfile_RotD50_d050_public_version.xlsx \
#   --filenames-csv /home/jupyter/Work/stampede3/Datasets/NGAWest2/NGAWest2_FilenamesAndMeta.csv \
#   --hdf5 /home/jupyter/Work/stampede3/Datasets/NGAWest2/NGAWest2_TimeSeriesOnly_byRSN_AT2_260115.hdf5 \
#   --out-prefix NGAWest2 \
#   --outdir out_process \
#   --limit 50 \
#   --compute-rotd \
#   --ml-missing drop


